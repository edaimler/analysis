
from datetime import datetime
from openpyxl import load_workbook, Workbook
import pandas
from pathlib import Path
import pyodbc
import pdb
from dateutil.relativedelta import relativedelta

database = 'GOHS Grants.accdb'
#database = 'TEST.accdb'

def report_numbers():
    current_month = datetime.now().month
    late_report_number = current_month + 1 % 12 
    number_list = []
    for n in range(1, 1+ late_report_number):
        number_list.append(n%12)

    number_string = "("
    for i in number_list:
        number_string = number_string + ''' DocNumber = 'R'''+str(i)+ '''' OR'''
    number_string = number_string.rstrip('OR')
    number_string = number_string + '))'
        
    return number_string



def get_late_reports():
    path = Path('F:\TSREG')
    #database = 'GOHS Grants.accdb'
    conn_filename = path / database
    conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)}; DBQ=' + str(conn_filename) + ';')
    cursor = conn.cursor()
    # In Access you need parenthesis when you have more than one join
    tuples = cursor.execute('''SELECT Grantees.Organization, Grants.Year, Grants.DocTitle, ProgressReports.DocNumber
        FROM (ProgressReports 
        INNER JOIN Grants ON ProgressReports.GrantID = Grants.GrantID)
        INNER JOIN Grantees ON Grants.GranteeID = Grantees.GranteeID
        WHERE (Status < 3
        AND ''' + report_numbers()).fetchall()  

    return tuples

get_late_reports()


def excel_reports():
    wb = Workbook() 
    current_month = datetime.now().strftime("%B") 
    title = current_month + ' Late Reports'
    dest_filename = title+'.xlsx'
    ws1 = wb.create_sheet(title = "Reports") 
    data = [list(row) for row in get_late_reports()]
    current_grant_num = ((datetime.now().month - 10)%12) + 1 
    latest_due_date = datetime(datetime.now().year, datetime.now().month, 20)
    column_titles = ["Organization", "Grant Year", "Grant Number", "Report Number", "Due Date"]
    ws1.append(column_titles)
    
    for row in data:
        report_num = int(row[-1][1:])
        due_date = latest_due_date - relativedelta(months= (current_grant_num - (report_num+1)))
        row.append(due_date)

        ws1.append(row)
    wb.save(dest_filename)

excel_reports()