from openpyxl import load_workbook, Workbook
from docx import Document
from docxcompose.composer import Composer
from docx import Document as Document_compose
from pathlib import Path
import pdb
import pyodbc

import urllib
from bs4 import BeautifulSoup
from selenium import webdriver
from urllib.parse import urljoin
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests.auth import HTTPBasicAuth

import claimsfunctions
import grantsfunctions
import codecs
import doc_elements
import final_reports
import form_elements
import functions
import progress_reports
import pandas
import queries
import urllib.request
import re
import time
import requests
        #"    "GA-2021-F.A.S.T. 405c M3DA-077",
        #
orgs = ["insert org" 
        ]

master_doc = Document_compose('annual_report_docs/blank.docx')
composer = Composer(master_doc)




def remove_row(table, row):
    tbl = table._tbl
    tr = row._tr
    tbl.remove(tr)

def parse_grant_name(grant_name):
    parsed = grant_name.split("-")

    document = {
    "docname": grant_name,
    "program": parsed[0],
    "classification": parsed[2],
    "year": parsed[1],
    "id": parsed[3]
    }
    
    return document
    

# fill in grant info here, then fill in activities below

        

gohsform = form_elements.gohsform
url = 'insert url'

#grant_name = orgs[1]


with requests.Session() as session:

        response = functions.login(session)

        for grant_name in orgs:

                word_doc = Document("AR_Appendix.docx")
                tables = word_doc.tables
                doc_grantee = tables[0].cell(0,1)			
                doc_grant_number = tables[0].cell(0,4)
                doc_grant_title = tables[0].cell(1,1)	
                doc_funding_source = tables[0].cell(1,4)
                doc_total_budget = tables[0].cell(2,1)			
                doc_total_expended = tables[0].cell(2,4)
                doc_project_desc = tables[0].cell(3,1)
                doc_state_targets = tables[0].cell(4,1)

                grant_list = grantsfunctions.get_grants(session, response, grant_name)
                final_list = final_reports.get_final_reports(session, response, grant_name)

                print(final_list)
                print(grant_list)

                if len(final_list) > 0:
                        final = final_list[0] 
                else: 
                        continue

                if len(grant_list) > 0:
                        grant = grant_list[0] 
                else: 
                        continue


                if final_reports.get_final_status(session, final) > 1: 
                        grant_status = grantsfunctions.get_grant_status(session, grant)

                        grant_desc = grantsfunctions.get_project_title_and_summary(session, grant)

                        grant_title = grant_desc[0]
                        grant_summary = grant_desc[1]

                        grant_budget = grantsfunctions.get_budget(session, grant)
                        budget_spent = final_reports.get_total_spent(session, final)

                        activities = final_reports.get_activities_and_results(session, final, grant_name)

                        funding_source = parse_grant_name(grant_name)["classification"]

                doc_grantee.text = grant_name		
                doc_grant_number.text  = grant_name
                doc_grant_title.text  = grant_title	
                doc_funding_source.text  = funding_source
                doc_total_budget.text  = grant_budget			
                doc_total_expended.text  = budget_spent
                doc_project_desc.text  = grant_summary
                doc_state_targets.text  = ''

                
                tables.pop(0)
                while len(activities) > 0:
                        for table in tables: 
                                if len(activities) < 1:
                                        table._element.getparent().remove(table._element)
                                        continue
                                
                                for row in table.rows: 
                                        if len(activities) < 1:
                                                remove_row(table, row)
                                                continue
                                        if row.cells[0].text == 'Activities Funded/Implemented': 
                                                continue
                                        activity = activities[0]
                                        row.cells[0].text = activity[0]
                                        row.cells[1].text = activity[1] 
                                        activities.pop(0) 
                


                word_doc.save('annual_report_docs/test.docx')

                doc2 = Document_compose('annual_report_docs/test.docx')
                composer.append(doc2)
                composer.save('annual_report_docs/GOHS Grants.docx')
