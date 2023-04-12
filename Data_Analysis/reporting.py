#from crypt import methods
from datetime import datetime, timedelta
from tkinter.tix import INTEGER
from traceback import StackSummary
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas
from pathlib import Path
import pyodbc
import pdb
import re

#print("login response page: ", BeautifulSoup(r.text).text)
'''
progressreports = functions.open_progress_reports(s, r)
#grant_list = load_workbook(filename = 'Grant List.xlsx')
#sheet_1 = grant_list['Grant List']
#grant_name = sheet_1.cell(row = 2, column = 2).value
response2 = functions.fetch(s, progressreports, url)
'''

filename = 'F:/Business/Evaluation/Auburn Prevention Project/FY2022/2022 Barrow School Survey/Barrow 2022 Survey.xlsx'
filename2 = 'F:/Business/Evaluation/Madison/Madison Pre-Survey 2022 Cleaned.xlsx'
filename3 = 'F:/Business/Evaluation/Stephens/Stephens Pre-Survey Cleaned.xlsx'

survey = pandas.read_excel(filename3, header = 1)
#survey = survey[survey['Progress'] > 50] #drop records with progress less than 50%

gender = survey.groupby('What is your gender?')
#school = survey.groupby('What school do you attend?')
#grade = survey.groupby('What grade are you in?')




def clean_numbers(answer_string):
    to_string = str(answer_string)
    cleaned = ""

    for i in to_string:
        if i.isdigit():
            cleaned = cleaned + i
        if i == '.':
            cleaned = cleaned + i

    result = None 

    try:
        result = float(cleaned)
    except:
        result = None 
    return result

def number_percent(answer_string):
    try:
        percent = len(str(clean_numbers(answer_string)))/len(str(answer_string))
    except: 
        percent = 0.0
    return percent

def clean_age(age):
    try:
        if age > 20: 
            return None 
        if age < 11 : 
            return None
    except:
        return None
    return age

def is_number_answer(series):
    numbers_per_answer = series.map(number_percent).dropna()
    all_answers = len(series)
    number_answers = len(series.map(clean_numbers).dropna())

    try:
        if number_answers > 0.5*all_answers and numbers_per_answer.mean() > .5:
            return True
        else:
            return False
    except:
        return False


survey['Cleaned Grades'] = survey['What grade are you in school?']
middle_school = survey[survey['What grade are you in school?'].map(clean_numbers) < 9]
#middle_school = middle_school[middle_school['Cleaned Grades'] > 5]

high_school = survey[survey['What grade are you in school?'].map(clean_numbers) > 8]
high_school = high_school[high_school['What grade are you in school?'].map(clean_numbers) < 13]

survey = high_school

executive_summary = ""

project_background = ""
methods = "This survey "

#results = [demographics, alcohol, marijuana, peer_use, peer_approval, parent_approval]

conclusions_and_recommendations = ""

class Question():
    def __init__(self, series):
        def sort(series):
            if is_number_answer(series):
                cleaned = series.map(clean_numbers).dropna().sort_values()
                if series.name == 'What is your age?':
                    cleaned.map(clean_age)
                return cleaned
            else: 
                return series
                
        self.series = sort(series)
        #self.series=series
        self.number_of_values = len(self.series.value_counts())
        #to get column name, use table[column].name
        self.groupby = ''
        self.caption = self.series.name + " (N = %s)" %self.series.count()
        try:
            self.label = self.series.name.replace(' ','').replace(',', '').replace('?', '').replace('(', '').replace(')', '').lower()[:15]
        except: 
            self.label = self.series.name.replace(' ','').replace(',', '').replace('?', '').replace('(', '').replace(')', '').lower()
        
        try:
            if is_number_answer(self.series):
                self.counts = self.series.value_counts(sort=True).sort_index()
            else:
                self.counts = self.series.value_counts(sort=True)

            self.percent = round(self.counts/(self.counts.sum())*100, 1)
            self.categories = tuple(self.counts.index)
        except: 
            print("create Question error: %s is not a series" %series)
    
    def make_table(self, save = False):
                response = 'Response'
                begin_table = r'''
                    \begin{table}[h]
                    \begin{tabular}{ l c c }
                    '''
                column_titles = '''
                    \\textbf{%s} & \\textbf{Count} & \\textbf{Percent of Total} \\\\
                    \hline
                    '''%response
                tuples = tuple(zip(self.categories, self.counts, self.percent))
                table_row_strings = []
                intro = begin_table + column_titles
                for (index, count_tuple) in enumerate(tuples): 
                    if index % 2 > 0:
                        row = '%s & %s & %s%s \\\\ \n' %(count_tuple + (r'\%',))
                    else: 
                        row = r'\rowcolor[gray]{0.95} ' + '%s & %s & %s%s \\\\ \n' %(count_tuple + (r'\%',))
                    table_row_strings.append(row)
                    intro = intro + row
                end_table = '''
                    \hline
                    \end{tabular}
                    \centering
                    \caption{%s}
                    \label{table:%s}
                    \hspace*{0pt}\hfill
                    \end{table}
                    ''' %(self.caption, self.label)
                    
                self.table = intro + end_table

                if save: 
                    with open('charts.txt', 'a') as output_file:  
                        output_file.write(self.table)

                return self.table

    def make_figure(self, xlabel, save = False):
        x_coords = []
        for i in [a for a in self.categories]: 
            try:
                x_coords.append(int(i))
            except: 
                x_coords.append(i)
        coordinates = tuple(zip(tuple(x_coords), self.counts))
        x_coords_string = '{' + str(x_coords)[1:-1] + '}'
        coordinates_string = '{' + str(coordinates)[1:-1] + '}'
        coordinates_string = coordinates_string.replace('),', ')')
        begin_figure = r'''\begin{figure}[h]
        \begin{center}
        \begin{tikzpicture}
        \begin{axis}[ybar,
        %ymin=0,
        %enlarge x limits=.25,
        x=-0.5cm,
        %bar width=1cm,
        ymin=0,
        %ymax=50,
        %legend style={at={(1.6, 0.6)}}, %Positions the Chart's Legend
        ylabel={Number of Respondents}, 
        xlabel ={ ''' + xlabel + r'''},
        symbolic x coords= ''' + x_coords_string + r''',
        xtick=data,
        x tick label style={rotate=0, anchor=north},
        nodes near coords, 
        ]
    
        \addplot+ [ybar, draw=herrick, fill=herrick, text=black] coordinates ''' + coordinates_string + r''';
        %\addplot+ [ybar, draw=odyssey, fill=odyssey] coordinates{(65 and Older, 62) (45 to 64,64) (25 to 44,40) (18 to 24, 20)};
        %\legend{Distraction-Prone, Distraction-Averse};
        
        \end{axis}
        \end{tikzpicture}
        \caption{''' + self.caption + r'''}
        \label{table:''' + self.label + r'''}
        \end{center}
        \end{figure}
        '''
        self.figure = begin_figure

        if save: 
            with open('charts.txt', 'a') as output_file:
               output_file.write(self.figure) 
          
        return begin_figure

class Table():
    def __init__(self, dataframe):
        self.dataframe = pandas.DataFrame(dataframe)
        #self.dataframe['What is your age?'] = self.dataframe['What is your age?'].map(clean_numbers).dropna()
        self.dataframe.fillna("No Response", inplace = True)
        self.series = {}
        for i in self.dataframe.columns: 
            self.series[i] = Question(self.dataframe[i])
        self.columns = self.dataframe.columns
        self.column = self.dataframe.columns[0]
        self.title = self.column 
        self.groupby = ''
        self.caption = self.groupby
        self.label = self.caption.replace(' ','').replace(',', '').replace('?', '').lower()

    def __getitem__(self, item):
        return self.series[item]

    def assign_column(self, column):
        try: 
            self.column = str(column)
            self.series = self.dataframe[column]
            self.title = self.column
            
        except:
            print("assign column error: %s is not a column of the dataframe" %column)
    
    def assign_title(self, title):
        try: 
            self.title = str(title)
        except:
            print("create title error: %s is not a string" %title)
        
    def assign_caption(self, caption):
        try:
            self.caption = str(caption)
            self.label = self.caption.replace(' ','').replace(',', '').replace('?', '').lower()
        except:
            print("create caption error: %s is not a string" %caption)
    
    def assign_group(self, column_name):
        try:
            self.group = self.dataframe.groupby(column_name)
        except:
            print("create group error: %s is not a column of the dataframe" %column_name)

    

    #for column in columns: 
        #assign_column(column)
        #table.assign_caption(column) 
        #table_string = ''
        #table_string = table_string + " \n " + " \n " + table.gen_table()


table = Table(survey)


#with open('charts.txt', 'w') as output_file:  
#    output_file.write(table_string)

first_columns = table.columns[10:13]
second_columns = table.columns[15:20]
third_columns = table.columns[32:-1]



str1 = 'About how many times during the past year have you seen or heard information that most students at your school do not drink alcohol?'

str2 = 'How many parents in your community do you think allow alcohol at parties thrown by their children?'

str3 = 'Think back over the last 30 days.  On how many days, if any, did you drink one or more drinks of an alcoholic beverage (Alcoholic beverages include beer, wine, wine coolers, malt beverages and liquor)? - Write the number of days (0 to 30): - Text'

str4 = 'Think back over the last 30 days.  On how many days, if any, did a typical student (peers, friends, or people your own age) in your school drink one or more drinks of an alcoholic beverage (Alcoholic beverages include beer, wine, wine coolers, malt beverages and liquor)? - Write the number of days (0 to 30): - Text'

str5 = 'Think back over the last 30 days.  On how many days, if any, did you drink five or more drinks of an alcoholic beverage within about 2 hours (Alcoholic beverages include beer, wine, wine coolers, malt beverages and liquor)? - Write the number of days (0 to 30): - Text'

str6 = 'Think back over the last 30 days.  On how many days, if any, did a typical student (peers, friends, or people your own age) in your school drink five or more drinks of an alcoholic beverage within about 2 hours (Alcoholic beverages include beer, wine, wine coolers, malt beverages and liquor)? - Write the number of days (0 to 30): - Text'

str7 = 'Think back over the last 30 days.  On how many days, if any, did you use marijuana (including smoking, edibles, or vaping)? - Write the number of days (0 to 30): - Text'

str8 = 'Think back over the last 30 days.  On how many days, if any, did a typical student (peers, friends, or people your own age) in your school use marijuana (including smoking, edibles, or vaping)? - Write the number of days (0 to 30): - Text'

for (i, column) in enumerate(first_columns): 
    table[column].label = table[column].label+str(i+1) 
    table[column].make_table(save = True)

for (i, column) in enumerate(second_columns): 
    table[column].label = table[column].label+str(i+1) 
    table[column].make_table(save = True)

table[str1].make_figure(xlabel = 'Number of Times Heard', save = True)

table[str2].make_table(save = True)

table[str3].make_figure(xlabel = 'Frequency of Alcohol Use Over Past 30 Days', save = True)

table[str4].make_figure(xlabel = 'Perceived Frequency of Peer Alcohol Use Over Past 30 Days', save = True)

#table[str5].make_figure(xlabel = 'Frequency of Binge Drinking Over Past 30 Days', save = True)

#table[str6].make_figure(xlabel = 'Perceived Frequency of Peer Binge Drinking Over Past 30 Days', save = True)

#table[str7].make_figure(xlabel = 'Frequency of Marijuana Use Over Past 30 Days', save = True)

#table[str8].make_figure(xlabel = 'Perceived Frequency of Peer Marijuana Use Over Past 30 Days', save = True)



for (i, column) in enumerate(third_columns): 
    table[column].label = table[column].label+str(i+1) 
    table[column].make_table(save = True)

test = table['What is your age?']

test2 = table['If you wanted to get some alcoholic beverages, how easy would it be for you to get it (Alcoholic beverages include beer, wine, wine coolers, malt beverages and liquor)?']

test3 = table['Think back over the last 30 days, how did you usually get the alcohol you drank? (please select all that apply) - Selected Choice']

pdb.set_trace()

with open('charts.txt', 'w') as output_file:
    output_file.write('') 