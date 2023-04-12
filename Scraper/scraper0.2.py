import urllib
from bs4 import BeautifulSoup
from selenium import webdriver
from urllib.parse import urljoin
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests.auth import HTTPBasicAuth

import codecs
import doc_elements
import form_elements
import functions
import navigation
import pandas
import urllib.request
import re
import time
import requests


url = 'insert url'
url2 = 'insert url 2'


gohsform = form_elements.gohsform


with requests.Session() as s:

    #================= LOGIN =================

    r = functions.login(s)

    time.sleep(10)

    # end login

    #=========================================
    #print("login response page: ", BeautifulSoup(r.text).text)
    '''
    progressreports = functions.open_progress_reports(s, r)
    #grant_list = load_workbook(filename = 'Grant List.xlsx')
    #sheet_1 = grant_list['Grant List']
    #grant_name = sheet_1.cell(row = 2, column = 2).value
    response2 = functions.fetch(s, progressreports, url)

    #print("getting form fields from final reports page")
    
    time.sleep(5)

    form2 = BeautifulSoup(response2).find(gohsform['tag'], id=gohsform['id'])
    formdata2 = functions.prep_form(response2, gohsform)

    for i in form_elements.pop_list:
        formdata2.pop(i)  # TODO REMOVE 'CLEAR' value from form

    progressurl = urljoin(progressreports.url, form2['action']) # get the URL of progressreports page
    progressreport = s.post(progressurl, data = formdata2) 
    time.sleep(10)
    '''
    orgs = ["Insert org list",
        ]
    '''
    for i in orgs:
        time.sleep(5)

        form3 = BeautifulSoup(response2).find(gohsform['tag'], id=gohsform['id'])
        formdata3 = functions.java_form(response2, gohsform, '', i)

        for j in form_elements.pop_list:
            formdata3.pop(j)  # TODO REMOVE 'CLEAR' value from form


        progressurl = urljoin(progressreports.url, form3['action']) # get the URL of progressreports page
        time.sleep(10)
        progressreport = s.post(progressurl, data = formdata3) 
        #progressreport = s.get(progressurl)

        proglist = functions.doc_list(progressreport.text)
        print('proglist is: ', proglist)
        
        for link in proglist:
            firstlink = functions.open_link(s, link)

            #print("first link is: ", BeautifulSoup(firstlink.text))

            mytable = BeautifulSoup(firstlink.text).find('table', id="ctl00_cphPageContent_wclDocumentForms_dgdDocumentMenu")

            frame = pandas.read_html(str(mytable))[0]

            doc_dict = functions.parse_doc_title(firstlink.text)

            print("parsed doc title is: ", doc_dict)

            navigation.save_file(doc_dict, frame, "name")
            navigation.save_file(doc_dict, frame, "namey_name")

            monthly = functions.open_id(s, firstlink.text, doc_elements.monthly)
            monthlytable = BeautifulSoup(monthly.text).find('table', id=re.compile('^ctl00_cphPageContent_tblPageTable'))
            monthlyframe = pandas.read_html(str(monthlytable))[0]
            navigation.save_file(doc_dict, monthlyframe, "monthly")

            narrative = functions.open_id(s, firstlink.text, doc_elements.narrative)
            narrativetable = BeautifulSoup(narrative.text).find('table', id=re.compile('^ctl00_cphPageContent_tblPageTable'))
            narrativeframe = pandas.read_html(str(narrativetable))[0]
            navigation.save_file(doc_dict, narrativeframe, "narrative")

            enforcement = functions.open_id(s, firstlink.text, doc_elements.enforcement)
            enforcementtable = BeautifulSoup(enforcement.text).find('table', id=re.compile('^ctl00_cphPageContent_tblPageTable'))
            enforcementframe = pandas.read_html(str(enforcementtable))[0]
            navigation.save_file(doc_dict, enforcementframe, "enforcement")

            milestone = functions.open_id(s, firstlink.text, doc_elements.milestone)
            milestonetable = BeautifulSoup(milestone.text).find('table', id=re.compile('^ctl00_cphPageContent_tblPageTable'))
            if milestonetable == None:
                milestonetable = BeautifulSoup(milestone.text).find('table', id="ctl00_cphPageContent_tblPageTable13113")
            if milestonetable == None:
                milestonetable = BeautifulSoup(milestone.text).find('table', id="ctl00_cphPageContent_tblPageTable13823")
            milestoneframe = pandas.read_html(str(milestonetable))[0]
            navigation.save_file(doc_dict, milestoneframe, "milestone")
            input_df = functions.get_milestone_inputs(milestone)
            navigation.save_file(doc_dict, input_df, "achieved_milestones")

            statistical = functions.open_id(s, firstlink.text, doc_elements.statistical)
            statisticaltable = BeautifulSoup(statistical.text).find('table', id=re.compile('^ctl00_cphPageContent_tblPageTable'))
            statisticalframe = pandas.read_html(str(statisticaltable))[0]
            navigation.save_file(doc_dict, statisticalframe, "statistical")
            

    '''
    claims = functions.open_claims(s, r.text)
    response3 = functions.fetch(s, claims.url)

    for i in orgs:
        time.sleep(5)

        form3 = BeautifulSoup(response3).find(gohsform['tag'], id=gohsform['id'])
        formdata3 = functions.java_form(response3, gohsform, '', i)

        for i in form_elements.pop_list:
            formdata3.pop(i)  # TODO REMOVE 'CLEAR' value from form


        claimurl = urljoin(claims.url, form3['action']) # get the URL of progressreports page
        claimreport = s.post(claimurl, data = formdata3) 
        time.sleep(10)

        claimlist = functions.doc_list(claimreport.text)

        for link in claimlist:
            secondlink = functions.open_link(s, link)

            #print("first link is: ", BeautifulSoup(secondlink.text))

            mytable2 = BeautifulSoup(secondlink.text).find('table', id="ctl00_cphPageContent_wclDocumentForms_dgdDocumentMenu")

            frame2 = pandas.read_html(str(mytable2))[0]

            doc_dict2 = functions.parse_doc_title(secondlink.text)

            print("parsed doc title is: ", doc_dict2)

            navigation.save_file(doc_dict2, frame2, "details")

            expense = functions.open_id(s, secondlink.text, doc_elements.expense)
            expensetable = BeautifulSoup(expense.text).find('table', id=re.compile('ctl00_cphPageContent_tblPageTable13868'))
            #print("expense table is: ", expensetable)
            expenseframe = pandas.read_html(str(expensetable))[0]
            navigation.save_file(doc_dict2, expenseframe, "expense")

