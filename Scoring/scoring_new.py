import score_functions
import score_navigation
import time
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

filepath = '2021Q3 Scores.xlsx'
wb = Workbook()
wb.save(filepath)
time.sleep(5)
grant_list = load_workbook(filename = filepath)

scores = grant_list['Sheet']

grants = score_navigation.late

for i in range(len(score_navigation.late)):
    print("i is: ", i)
    print("Grant is: ", grants[i])
    doc_info = score_functions.parse_doc_title(grants[i])
    print("organization is: ", doc_info['organization'])
    scores.cell(row = i+2, column = 1).value = doc_info['organization']
    scores.cell(row = i+2, column = 2).value = doc_info['grantname']
    scores.cell(row = i+2, column = 3).value = doc_info['program']
    scores.cell(row = i+2, column = 4).value = doc_info['classification']
    scores.cell(row = i+2, column = 5).value = score_navigation.goals(grants[i], 1)
    scores.cell(row = i+2, column = 6).value = score_navigation.milestones(grants[i], 1)
    scores.cell(row = i+2, column = 7).value = score_navigation.spending(grants[i], 1)
    scores.cell(row = i+2, column = 8).value = score_navigation.quality(grants[i], 1)
    scores.cell(row = i+2, column = 9).value = score_navigation.timeliness(grants[i], 1)
    weight = score_navigation.weight_score(grants[i], 1)
    scores.cell(row = i+2, column = 10).value = weight
    scores.cell(row = i+2, column = 11).value = score_navigation.grade_score(weight)
    grant_list.save(filepath)
