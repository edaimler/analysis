

import urllib
import datetime
import re
import urllib.request
import codecs
import time
import requests
#import bs
from datetime import timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests.auth import HTTPBasicAuth






    #=========================================

def grant_names():
    grant_list = load_workbook(filename = '2021 Q1Q2 All Data.xlsx')
    scores = grant_list['Scores']
    name_list = []
    for cell in scores['C']:
        if(cell.value is not None):
            name_list.append(cell.value)
    grant_list.close()
    name_list.pop(0)
    
    return name_list

def scoring(grant_name, report_list):
    grant_list = load_workbook(filename = '2021 Q1Q2 All Data.xlsx')
    scores = grant_list['Scores']
    timeliness_scores = []
    quality_scores = []
    for cell in scores['C']:
        if cell.value is not None:
            if grant_name in cell.value:
                for report in report_list: #report_list is the list of row numbers for a grant's progress reports
                    days_late = grant_list['Progress Reports']['K'][report-1].value
                    print("scoring days late is: ", days_late)
                    if grant_list['Progress Reports']['H'][report-1].value is not None:
                        quality_scores.append(grant_list['Progress Reports']['H'][report-1].value)
                    if days_late is not None:
                        if days_late == 0:
                            timeliness_scores.append(1)
                        elif days_late > 0 and days_late <= 90:
                            timeliness_scores.append(.5)
                        elif days_late > 90:
                            timeliness_scores.append(.25)
                        else:
                            timeliness_scores.append(0)
                    else:
                        timeliness_scores.append(0)
                print("timeliness_scores are: ", timeliness_scores)
                if len(timeliness_scores):
                    scores['K'][cell.row-1].value = sum(timeliness_scores)/len(timeliness_scores)
                if len(quality_scores):
                    scores['J'][cell.row-1].value = sum(quality_scores)/len(quality_scores)
    grant_list.save('2021 Q1Q2 All Data.xlsx')
    grant_list.close()

def final_report(grant_name):
    grant_list = load_workbook(filename = '2021 Q1Q2 All Data.xlsx')
    scores = grant_list['Scores']
    final = grant_list['Final ReportsQ1']
    report_name = grant_name+"-R6"
    goals_count = []
    milestones_list = []
    all_milestones = []
    spend = 0
    for cell in final['C']:
        if(cell.value is not None):
            if report_name in cell.value:
                #print("found value {} at row: {} and column: {}.  In cell {}".format(cell.value,cell.row,cell.column,cell))
                #print("cell row is: ", cell.row)
                total_milestones = final['J'][cell.row-1].value.strip('][').split(',')
                reached_milestones = final['K'][cell.row-1].value.strip('][').split(',')
                print("total milestones are: ", total_milestones)
                print("reached milestones are: ", reached_milestones)
                if total_milestones is not None and reached_milestones is not None:
                    if len(total_milestones) == len(reached_milestones):
                        for i in range(len(total_milestones)):
                            all_milestones.append(int(total_milestones[i]))
                            if (int(reached_milestones[i]) < int(total_milestones[i])):
                                print("failed total milestone is: ", total_milestones[i])
                                print("failed reached milestone is: ", reached_milestones[i])
                                goals_count.append(0)
                                milestones_list.append(int(reached_milestones[i]))
                            else:
                                print("i is: ", i)
                                print("total milestone is: ", total_milestones[i])
                                print("reached milestone is: ", reached_milestones[i])
                                goals_count.append(1)
                                milestones_list.append(int(total_milestones[i]))

                if (final['O'][cell.row-1].value is not None and final['P'][cell.row-1].value is not None):
                    spend = min(float(final['P'][cell.row-1].value)/float(final['O'][cell.row-1].value),1.0)

    for scores_cell in scores['C']:
        if(scores_cell.value is not None):
            if grant_name in scores_cell.value:
                if len(goals_count) > 0:
                    print("goals count is: ", goals_count)
                    print("milestones list is: ", milestones_list)
                    print("all milestones is: ", all_milestones)
                    scores['G'][scores_cell.row-1].value = sum(goals_count)/len(goals_count)
                    scores['H'][scores_cell.row-1].value = sum(milestones_list)/sum(all_milestones)
                    scores['I'][scores_cell.row-1].value = spend
    grant_list.save('2021 Q1Q2 All Data.xlsx')
    grant_list.close()


                            
                
    grant_list.close()
    return cell.row
    grant_list.close()

def progress_report(grant_name):
    grant_list = load_workbook(filename = '2021 Q1Q2 All Data.xlsx')
    progress = grant_list['Progress Reports']
    report_rows = []
    for cell in progress['C']:
        if(cell.value is not None):
            if grant_name in cell.value:
                print("found value {} at row: {} and column: {}.  In cell {}".format(cell.value,cell.row,cell.column,cell))
                print("cell row is: ", cell.row)
                if progress['G'][cell.row-1].value is not None:
                    if cell.value.endswith("-R1"):
                        duedate = datetime.datetime(2020, 11, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R1 due date: ", progress['F'][cell.row-1].value)
                        print("R1 last modified: ", progress['G'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R2" in cell.value:
                        duedate = datetime.datetime(2020, 12, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R2 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R3" in cell.value:
                        duedate = datetime.datetime(2021, 1, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R3 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R4" in cell.value:
                        duedate = datetime.datetime(2021, 2, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R4 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R5" in cell.value:
                        duedate = datetime.datetime(2021, 3, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R5 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R6" in cell.value:
                        duedate = datetime.datetime(2021, 4, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R6 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R7" in cell.value:
                        duedate = datetime.datetime(2021, 5, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R7 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R8" in cell.value:
                        duedate = datetime.datetime(2021, 6, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R8 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R9" in cell.value:
                        duedate = datetime.datetime(2021, 7, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R9 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R10" in cell.value:
                        duedate = datetime.datetime(2021, 8, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R10 due date: ", progress['F'][cell.row-1].value)
                        print("R10 type: ", type(progress['F'][cell.row-1].value))
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R11" in cell.value:
                        duedate = datetime.datetime(2021, 9, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R11 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                    if "-R12" in cell.value:
                        duedate = datetime.datetime(2021, 10, 20)
                        progress['F'][cell.row-1].value = duedate
                        print("R12 due date: ", progress['F'][cell.row-1].value)
                        if (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days > timedelta(days = 0).days:
                            dayslate = (progress['G'][cell.row-1].value - progress['F'][cell.row-1].value).days
                        else :
                            dayslate = 0
                        progress['K'][cell.row-1].value = dayslate
                        print("days late is: ", dayslate)
                
                report_rows.append(cell.row)

    grant_list.save('2021 Q1Q2 All Data.xlsx')
    grant_list.close()
    return report_rows


def compute_weighted(grant_name):
    grant_list = load_workbook(filename = '2021 Q1Q2 All Data.xlsx')
    scores = grant_list['Scores']
    weighted_score = scores['U']
    grades = scores['V']
    for cell in scores['C']:
        if cell.value is not None:
            if grant_name in cell.value: 
                if scores['G'][cell.row-1].value is not None and scores['H'][cell.row-1].value is not None and scores['I'][cell.row-1].value is not None and scores['J'][cell.row-1].value is not None and scores['K'][cell.row-1].value is not None:
                    goals = float(scores['G'][cell.row-1].value)
                    milestones = float(scores['H'][cell.row-1].value)
                    spend = float(scores['I'][cell.row-1].value)
                    quality = float(scores['J'][cell.row-1].value)
                    timeliness = float(scores['K'][cell.row-1].value)
                    score = (milestones)*.5 + (spend*.25) + (quality*.6 + timeliness*.4)*.25
                    weighted_score[cell.row-1].value = score
                    if score > .954:
                        grades[cell.row-1].value = "A+"
                    elif score >= .895:
                        grades[cell.row-1].value = "A"
                    elif score >= .795:
                        grades[cell.row-1].value = "B"
                    elif score >= .695:
                        grades[cell.row-1].value = "C"
                    else:
                        grades[cell.row-1].value = "D"
    grant_list.save('2021 Q1Q2 All Data.xlsx')
    grant_list.close()
'''
grant_name = "SADD-2020-402 TSP-003"
print("grant name is: ", grant_name)
report_list = progress_report(grant_name)
print("name list is: ", grant_names())
print("testing: ", report_list)
scoring(grant_name, report_list)
final_report(grant_name)
compute_weighted(grant_name)

'''
for grant_name in grant_names():
    print("grant name is: ", grant_name)
    report_list = progress_report(grant_name)
    print("name list is: ", grant_names())
    print("testing: ", report_list)
    scoring(grant_name, report_list)
    final_report(grant_name)
    compute_weighted(grant_name)
